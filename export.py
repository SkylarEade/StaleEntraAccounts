from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment
import os

def export_to_xlsx(data, output_path, sheet_name):
    if not data:
        print("No data to export.")
        return
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    headers = list(data[0].keys())
    ws.append(headers)

    for row in data:
        ws.append([row.get(h, "") for h in headers])

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    table_ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
    table = Table(displayName="ExportTable", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    wb.save(output_path)
    print(f"Exported {len(data)} rows to {output_path}")